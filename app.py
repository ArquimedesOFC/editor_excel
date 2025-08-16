from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import os

app = Flask(__name__)
CORS(app)

EXCEL_PATH = "Checklist_Sala_de_Aula_Sim_Nao.xlsx"

# Inicializar Excel se não existir
if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append(["Pergunta", "Sim", "Não"])
    perguntas = [
        "Cadeiras estão alinhadas e organizadas?",
        "Mesas limpas e organizadas?",
        "Quadro apagado e limpo?",
        "Lixo esvaziado corretamente?",
        "Materiais guardados?",
        "Ventilação adequada?",
        "Iluminação adequada?"
    ]
    for p in perguntas:
        ws.append([p, "", ""])
    wb.save(EXCEL_PATH)

@app.route("/")
def index():
    perguntas = [
        "Cadeiras estão alinhadas e organizadas?",
        "Mesas limpas e organizadas?",
        "Quadro apagado e limpo?",
        "Lixo esvaziado corretamente?",
        "Materiais guardados?",
        "Ventilação adequada?",
        "Iluminação adequada?"
    ]
    return render_template("index.html", perguntas=perguntas, sim_count=0, nao_count=0)

@app.route("/salvar", methods=["POST"])
def salvar():
    dados = request.json
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        # Mapear perguntas para linhas
        mapa = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            mapa[row[0]] = i

        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for pergunta, resposta in dados.items():
            linha = mapa.get(pergunta)
            if linha:
                if resposta == "Sim":
                    ws.cell(row=linha, column=2, value="Sim").fill = verde
                    ws.cell(row=linha, column=3, value="")
                elif resposta == "Não":
                    ws.cell(row=linha, column=3, value="Não").fill = vermelho
                    ws.cell(row=linha, column=2, value="")

        wb.save(EXCEL_PATH)
        return jsonify({"status": "sucesso"})
    except Exception as e:
        return jsonify({"status": "erro", "mensagem": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True)
